#!/usr/bin/env python3
"""
Import Basaltemperatur data from Excel into Supabase.
Uses Zyklustag (cycle day) resets to 1 to identify period start dates.
Period days = first 5 days of each cycle (Zyklustag 1-5).
"""

import openpyxl
import requests
import json
import sys
import os

# --- Configuration ---
SUPABASE_URL = "https://scohibllvlqujmvtuamv.supabase.co"
# Service role key for admin access (bypasses RLS)
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")

# User to import for
USER_EMAIL = "hi@sarahhoffmann.eu"

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Basaltemperatur Kopie.xlsx")

PERIOD_DAYS = 5  # Mark first N days of each cycle as period

def get_user_id(email: str) -> str:
    """Look up user ID by email via Supabase Auth admin API."""
    url = f"{SUPABASE_URL}/auth/v1/admin/users"
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    }
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    users = resp.json().get("users", [])
    for user in users:
        if user.get("email") == email:
            return user["id"]
    raise ValueError(f"User {email} not found")


def read_excel(path: str):
    """Read Excel file and return temperature entries and period entries.
    Column D contains 'Ja'/'Nein' for period. Must use data_only=True to read computed values."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    temp_entries = []
    period_entries = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        date_val = row[1].value
        temp = row[2].value
        periode = row[3].value  # Column D: "Ja" or "Nein"

        if date_val is None:
            continue

        date_str = date_val.strftime("%Y-%m-%d")

        # Temperature entry (only if we have a temperature)
        if temp is not None:
            temp_entries.append({
                "date": date_str,
                "temperature": float(temp),
            })

        # Period from Column D
        if isinstance(periode, str) and periode.strip().lower() == "ja":
            period_entries.append({
                "date": date_str,
                "flow_intensity": "medium",
            })

    return temp_entries, period_entries


def upload_entries(entries: list, table: str, user_id: str):
    """Upload entries to Supabase table."""
    if not entries:
        print(f"  No {table} entries to upload.")
        return

    # Add user_id to each entry
    for entry in entries:
        entry["user_id"] = user_id

    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }

    # Upload in batches of 50
    batch_size = 50
    total = len(entries)
    for i in range(0, total, batch_size):
        batch = entries[i:i + batch_size]
        resp = requests.post(url, headers=headers, json=batch)
        if resp.status_code >= 300:
            print(f"  ERROR batch {i // batch_size + 1}: {resp.status_code} {resp.text}")
            return False
        print(f"  Batch {i // batch_size + 1}/{(total + batch_size - 1) // batch_size} uploaded ({len(batch)} rows)")

    print(f"  ✅ {total} {table} entries uploaded successfully.")
    return True


def main():
    if not SUPABASE_SERVICE_KEY:
        print("ERROR: Set SUPABASE_SERVICE_KEY environment variable")
        print("  export SUPABASE_SERVICE_KEY='your-service-role-key'")
        sys.exit(1)

    print(f"📂 Reading Excel: {EXCEL_PATH}")
    temp_entries, period_entries = read_excel(EXCEL_PATH)
    print(f"  Found {len(temp_entries)} temperature entries")
    print(f"  Found {len(period_entries)} period entries")

    print(f"\n🔍 Looking up user: {USER_EMAIL}")
    user_id = get_user_id(USER_EMAIL)
    print(f"  User ID: {user_id}")

    print(f"\n📤 Uploading temperature entries...")
    upload_entries(temp_entries, "temperature_entries", user_id)

    print(f"\n📤 Uploading period entries...")
    upload_entries(period_entries, "period_entries", user_id)

    print(f"\n🎉 Import complete!")


if __name__ == "__main__":
    main()
